#----------------------
# Import Library & Workbook
#----------------------
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from datetime import datetime
import os
import shutil
from kmz_module import getAllHP, getAllFAT

# Process each file
def hpdbCheck(raw_file_path, kmz_file_path, cluster, checking_date, checking_time):
    print(kmz_file_path)
    start_time = time.time()
    file = os.path.basename(raw_file_path)
    homepass_long_lat = getAllHP(kmz_file_path)
    fat_long_lat = getAllFAT(kmz_file_path)

    # Define the directory paths
    output_dir = 'Output'
    summary_dir = f"Summary"
    summary_file_path = os.path.join(summary_dir, f"Checking Summary.xlsx")

    # Construct the full file paths
    output_file_path = os.path.join(output_dir, f"output_{file}")

    print("Processing = " + file)

    # Load the "kodepos.xlsx" Excel file into a pandas DataFrame
    kodepos_df = pd.read_excel('Reference/ZIP_Code.xlsx')

    # Load the "Mobile_Region_Cluster.xlsx" Excel file into a pandas DataFrame
    mobile_df = pd.read_excel('Reference/Mobile_Region_Cluster.xlsx')

    print(raw_file_path)

    # Load the file into a pandas DataFrame
    hpdb_df = pd.read_excel(raw_file_path,dtype={"HOMEPASS_ID" : str, "RT" : str, "RW" : str}, sheet_name='HPDB_Excel')

    # Load City_Code.xlsx into a DataFrame
    city_code_df = pd.read_excel('Reference/City_Code.xlsx')

    #----------------------
    # City Code Lookup
    #----------------------
    # Merge the two DataFrames on the 'CITY' column
    merged_df = pd.merge(hpdb_df, city_code_df[['CITY', 'CITY_CODE']], on='CITY', how='left')

    # Update CITY_CODE in HPDB SAMPEL.xlsx with the values from merged_df
    hpdb_df['CITY_CODE'] = merged_df['CITY_CODE_y']

    #----------------------
    # Mobile Region and Cluster Lookup
    #----------------------
    # Create a copy of the CITY column in hpdb_df
    hpdb_df['CITY_original'] = hpdb_df['CITY']

    # Convert CITY columns to lowercase for case-insensitive comparison
    hpdb_df['CITY'] = hpdb_df['CITY'].str.lower()
    mobile_df['CITY'] = mobile_df['CITY'].str.lower()

    # Merge the DataFrames on 'CITY' column
    merged_df = pd.merge(hpdb_df, mobile_df, on='CITY', how='left', suffixes=('', '_MOBILE'))

    # Replace the values in MOBILE_REGION and MOBILE_CLUSTER with the corresponding values from REGION and CLUSTER
    merged_df['MOBILE_REGION'] = merged_df['REGION MOBILE'].combine_first(merged_df['MOBILE_REGION'])
    merged_df['MOBILE_CLUSTER'] = merged_df['CLUSTER MOBILE'].combine_first(merged_df['MOBILE_CLUSTER'])

    # Drop the extra 'REGION' and 'CLUSTER' columns
    merged_df.drop(['REGION MOBILE', 'CLUSTER MOBILE', 'PROVINCE', 'CITY.1'], axis=1, inplace=True)

    # Replace CITY column with original values
    merged_df['CITY'] = merged_df['CITY_original']

    # Drop the 'CITY_original' column
    merged_df.drop('CITY_original', axis=1, inplace=True)

    #----------------------
    # Null Filling
    #----------------------
    # Replace null values with '-'
    hpdb_filled_df = merged_df.fillna('-')

    # Remove double spaces from all columns
    hpdb_filled_df = hpdb_filled_df.replace(r'\s+', ' ', regex=True)

    # Remove '|' characters from PROJECT_NAME column
    hpdb_filled_df['PROJECT_NAME'] = hpdb_filled_df['PROJECT_NAME'].str.replace('|', '')

    # Create new column based on concatenated columns
    hpdb_filled_df['ADDRESS'] = hpdb_filled_df[['PREFIX_ADDRESS', 'STREET_NAME', 'HOUSE_NUMBER', 'BLOCK', 'FLOOR', 'RT', 'RW']].apply(lambda x: ' '.join(x.astype(str)), axis=1)

    # Save the filled DataFrame to a new Excel file
    hpdb_filled_df.to_excel('Temp/temp.xlsx', index=False)

    # Load HPDB SAMPEL.xlsx into a DataFrame
    hpdb_df = hpdb_filled_df

    wb = load_workbook('Temp/temp.xlsx')
    ws = wb.active

    #----------------------
    # List and Variables
    #----------------------
    # Initialize a PatternFill object for the red color
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Get the current year
    current_year = datetime.now().year

    # Define the lists of allowed values
    acquisition_class_values = ["HOME", "HOME - BIZ", "BIZ - HOME", "BIZ"]
    building_type_values = ['PERUMAHAN', 'RUKO', 'FASUM']
    ownership_values = ['PARTNERSHIP-LN', 'PARTNERSHIP-IF', 'PARTNERSHIP-TBG', 'OWN BUILT']
    vendor_name_values = ['LINKNET', 'IFORTE', 'TBG']
    prefix_address_values = ['JL.', 'GG.']

    # Define the lists of similar columns
    coordinate_col = ['FDT_LONGITUDE', 'FAT_LONGITUDE', 'BUILDING_LONGITUDE', 'FDT_LATITUDE', 'FAT_LATITUDE', 'BUILDING_LATITUDE']
    rt_rw_col = ['RT','RW']
    kmz_col = ['Pole to FAT', 'Pole to FDT', 'HP to pole 35m', 'Coordinate HP to pole 35m', 'HP to FAT 150m', 'Coordinate HP to FAT 150m',
                "Pole not in Distribution and Sling", "Coordinate Pole not in Distribution and Sling"]
    log_col = ['Cluster ID', 'Checking date', 'Checking Time', "Status"]
    columns_to_keep = [
        'ACQUISITION_CLASS',
        'ACQUISITION_TIER',
        'BUILDING_TYPE',
        'OWNERSHIP',
        'VENDOR_NAME',
        'ZIP_CODE',
        'REGION',
        'CITY',
        'CITY_CODE',
        'DISTRICT',
        'SUB_DISTRICT',
        'FAT_CODE',
        'FAT_LONGITUDE',
        'FAT_LATITUDE',
        'BUILDING_LATITUDE',
        'BUILDING_LONGITUDE',
        'HOMEPASS_ID',
        'MOBILE_REGION',
        'MOBILE_CLUSTER',
        'CITY_GROUP'
    ]

    # Convert the allowed values to lowercase
    acquisition_class_values_lower = [value.lower() for value in acquisition_class_values]
    building_type_values_lower = [value.lower() for value in building_type_values]
    ownership_values_lower = [value.lower() for value in ownership_values]
    vendor_name_values_lower = [value.lower() for value in vendor_name_values]
    prefix_address_values_lower = [value.lower() for value in prefix_address_values]

    # Define the regular expression patterns for longitude and latitude
    longitude_pattern = r'^\d{1,3}\.\d{6}$'
    latitude_pattern = r'^-?\d{1,2}\.\d{6}$'

    # List to store row indices
    rows_to_delete = []
    red_columns = []

    #----------------------
    # Data Converting
    #----------------------
    # Convert the datetime columns to datetime objects
    hpdb_df['PARTNER_RFS_DATE'] = pd.to_datetime(hpdb_df['PARTNER_RFS_DATE'], format='%m/%d/%Y', errors='coerce')
    hpdb_df['RFS_DATE'] = pd.to_datetime(hpdb_df['RFS_DATE'], format='%m/%d/%Y', errors='coerce')

    # Convert string columns to lowercase
    hpdb_df = hpdb_df.map(lambda x: x.lower() if isinstance(x, str) else x)
    kodepos_df = kodepos_df.map(lambda x: str(x).lower() if isinstance(x, str) else x)

    #-------------------------
    # LONGITUDE and LATITUDE
    #-------------------------
    # Replace all commas with dots and truncate decimals to 6 digits
    for col in coordinate_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                value = str(cell.value).replace(',', '.')  # Replace commas with dots
                if '.' in value:
                    integer_part, decimal_part = value.split('.')
                    decimal_part = decimal_part.ljust(6, '0')[:6]  # Ensure decimal part has exactly 6 digits
                    value = f"{integer_part}.{decimal_part}"
                cell.value = value
                
    for index, row in hpdb_df.iterrows():
        build_longitude = row['BUILDING_LONGITUDE']
        build_latitude = row['BUILDING_LATITUDE']
        fat_longitude = row['FAT_LONGITUDE']
        fat_latitude = row['FAT_LATITUDE']
        if pd.notnull(build_longitude) and pd.notnull(build_latitude):
            # Ubah nilai longitude dan latitude menjadi tuple
            coordinate_tuple = [float(build_longitude), float(build_latitude)]

            if not str(coordinate_tuple) in homepass_long_lat["Coordinates"].astype(str).to_list():
                # Tambahkan logika untuk menandai jika nilai tidak ditemukan
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LONGITUDE') + 1).fill = red_fill
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LATITUDE') + 1).fill = red_fill

        if pd.notnull(fat_longitude) and pd.notnull(fat_latitude):
            # Ubah nilai longitude dan latitude menjadi tuple
            coordinate_tuple = [float(fat_longitude), float(fat_latitude)]

            if not str(coordinate_tuple) in fat_long_lat["Coordinates"].astype(str).to_list():
                # Tambahkan logika untuk menandai jika nilai tidak ditemukan
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LONGITUDE') + 1).fill = red_fill
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LATITUDE') + 1).fill = red_fill

    # Fill all existing rows with red color for RT and RW columns
    for col in rt_rw_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                if cell.value == "-":
                    for cell in row:
                        cell.fill = red_fill
    
    #----------------------
    # Zip Code Checking
    #----------------------
    # Iterate over each row in the "HPDB SAMPEL" DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if there is a matching row in the "kodepos" DataFrame
        match = kodepos_df[
            (kodepos_df['REGION'] == row['REGION']) &
            # (kodepos_df['CITY'] == row['CITY']) &
            (kodepos_df['DISTRICT'] == row['DISTRICT']) &
            (kodepos_df['SUB_DISTRICT'] == row['SUB_DISTRICT']) &
            (kodepos_df['ZIP_CODE'].astype(str) == str(row['ZIP_CODE']))
        ]
        # If a match is found, mark the row in the Excel worksheet for coloring
        if match.empty:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('REGION') + 1, max_col=hpdb_df.columns.get_loc('ZIP_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # ACQUISITION_CLASS
    #----------------------
    # Fill all existing rows with red color if ACQUISITION_CLASS is not in allowed_values
    for index, row in hpdb_df.iterrows():
        if row['ACQUISITION_CLASS'].lower() not in acquisition_class_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # BUILDING_TYPE
    #----------------------
    # Iterate over each row in the "HPDB SAMPEL" DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if BUILDING_TYPE is one of the allowed values
        if row['BUILDING_TYPE'] not in building_type_values_lower:
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1, max_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # ACQUISITION_TIER
    #----------------------
    # Iterate over each row in the "HPDB SAMPEL" DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if BUILDING_TYPE is one of the allowed values
        if row['BUILDING_TYPE'] == "ruko":
            #
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1):
                for c in cell:
                    c.value = "HOME - BIZ"

            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1):
                for c in cell:
                    c.value = "BIZ"
        else:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1):
                for c in cell:
                    c.value = "HOME"

    #----------------------
    # OWNERSHIP
    #----------------------
    # Check if any of the ownership values is in the OWNERSHIP field
    for index, row in hpdb_df.iterrows():
        found = False
        for value in ownership_values_lower:
            if value in row['OWNERSHIP']:
                found = True
                break
        if not found:
            # Fill the cell with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1, max_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1):
                for c in cell:
                    c.fill = red_fill
        elif row['OWNERSHIP'] != 'OWN BUILT' and row['VENDOR_NAME'] not in vendor_name_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1, max_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # PREFIX_ADDRESS
    #----------------------
    for index, row in hpdb_df.iterrows():
        if row['PREFIX_ADDRESS'] not in prefix_address_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1):
                for c in cell:
                    c.fill = red_fill

    #--------------------------------------------------------
    # CLUSTER_NAME, CLUSTER_CODE, PROJECT_NAME, CITY_CODE
    #--------------------------------------------------------
    # Iterate over each row in the DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if the CLUSTER_NAME column exceeds 100 characters
        if len(str(row['CLUSTER_NAME'])) > 100 or row['CLUSTER_NAME'] == "-":
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the CLUSTER_CODE column exceeds 20 characters
        if len(str(row['CLUSTER_CODE'])) > 20 or row['CLUSTER_CODE'] == "-":
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the length of PROJECT_NAME exceeds 100 characters
        if len(row['PROJECT_NAME']) > 100 or row['PROJECT_NAME'] == "-":
            # Fill the entire row with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1, max_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the length of PROJECT_NAME exceeds 100 characters
        if row['CITY_CODE'] == "-":
            # Fill the entire row with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CITY_CODE') + 1, max_col=hpdb_df.columns.get_loc('CITY_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

    #-------------
    # HOMEPASS_ID
    #-------------
    # Get the duplicated rows in the "HOMEPASS_ID" column
    duplicates = hpdb_df["HOMEPASS_ID"].duplicated()

    # Fill the duplicated rows with red color
    for index, duplicate in duplicates.items():
        if duplicate:
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('HOMEPASS_ID') + 1).fill = red_fill

    # Iterate over each row in the DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if the PARTNER_RFS_DATE column is NULL or has an incorrect format
        if pd.isnull(row['PARTNER_RFS_DATE']) or \
            not isinstance(row['PARTNER_RFS_DATE'], pd.Timestamp) or \
            row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y') != row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y'):
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('PARTNER_RFS_DATE') + 1).fill = red_fill

        # Check if the RFS_DATE column is NULL or has an incorrect format
        if pd.isnull(row['RFS_DATE']) or \
            not isinstance(row['RFS_DATE'], pd.Timestamp) or \
            row['RFS_DATE'].strftime('%m/%d/%Y') != row['RFS_DATE'].strftime('%m/%d/%Y'):
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('RFS_DATE') + 1).fill = red_fill

    #-----------------------------
    # Address Duplicate Checking
    #-----------------------------
    duplicates = hpdb_filled_df["ADDRESS"].duplicated()

    # Fill the duplicated rows with red color
    for index, duplicate in duplicates.items():
        if duplicate:
            for row in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('RW') + 1):
                for cell in row:
                    cell.fill = red_fill

    # Drop the "ADDRESS" column
    ws.delete_cols(hpdb_filled_df.columns.get_loc("ADDRESS") + 1, amount=1)

    #-------------
    # Save
    #-------------
    # Save the modified Excel workbook
    wb.save(output_file_path)

    #-----------------------------
    # Summary
    #-----------------------------
    # Load the Excel workbook
    wb = load_workbook(output_file_path)
    ws = wb.active

    # Iterate over each row starting from the second row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        # Check if any cell in the row has red fill
        has_red_fill = any([cell.fill == red_fill for cell in row])

        # If no red fill, add row index to rows_to_delete list
        if not has_red_fill:
            rows_to_delete.append(row[0].row)

    # Delete rows in reverse order to avoid shifting rows
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)

    # Iterate over each cell in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Skip certain columns
            if ws.cell(row=1, column=cell.column).value in ['FAT_CODE', 'FAT_LONGITUDE', 'FAT_LATITUDE', 'BUILDING_LATITUDE', 'BUILDING_LONGITUDE', 'HOMEPASS_ID']:
                continue

            # Check if cell has red fill
            if cell.fill == red_fill:
                cell.value = 'REVISE'
            else:
                cell.value = 'OK'

    wb.save('Temp/HPDB Summary.xlsx')

    # Create a DataFrame with column names from column_names
    hpdb_summary_df = pd.read_excel('Temp/HPDB Summary.xlsx')
    kmz_summary_df = pd.DataFrame(columns=kmz_col)

    # Menghapus kolom yang tidak ada dalam list
    columns_to_drop = set(hpdb_summary_df.columns) - set(columns_to_keep)
    hpdb_summary_df.drop(columns=columns_to_drop, inplace=True)

    hpdb_summary_num_rows = hpdb_summary_df.shape[0]

    # Create a new row with "-" values
    new_row = pd.DataFrame([["-"] * len(kmz_summary_df.columns)], columns=kmz_summary_df.columns)

    # Append the new row to kmz_df num_rows times
    for _ in range(hpdb_summary_num_rows):
        kmz_summary_df = kmz_summary_df._append(new_row, ignore_index=True)

    # Create a DataFrame with the log_columns and the new row
    log_summary_df = pd.DataFrame(columns=log_col)
    log_summary_df.loc[0] = [cluster, checking_date, checking_time, "REVISE"]

    # Repeat appending the new row to log_summary_df for hpdb_summary_num_rows times
    for _ in range(hpdb_summary_num_rows - 1):
        log_summary_df = log_summary_df._append(log_summary_df.iloc[0], ignore_index=True)

    summary_df = pd.concat([log_summary_df, kmz_summary_df, hpdb_summary_df], axis=1, ignore_index=False)

    # Create temp_master.xlsx with headers from master_temp_df if it doesn't exist
    if not os.path.exists(summary_file_path):
        summary_df.to_excel(summary_file_path, index=False, engine='openpyxl')
    
    else:
        with pd.ExcelWriter(summary_file_path, 'openpyxl', mode='a',  if_sheet_exists="overlay") as writer:
            # fix line
            reader = pd.read_excel(summary_file_path, sheet_name="Sheet1")
            summary_df.to_excel(writer, sheet_name="Sheet1", index=False, engine='openpyxl', header=False, startrow=len(reader)+1)

    #-------------
    # FInishing
    #-------------
    # Remove the raw file
    # os.remove(raw_file_path)
            
    end_time = time.time()
    execution_time = end_time - start_time
    print(execution_time)

    print("Done")