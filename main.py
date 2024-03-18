from datetime import datetime
import os
import random
import pandas as pd
import threading
import time
import concurrent.futures
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from datetime import datetime
import shutil
import simplekml
from fastkml import kml
from lxml import html
from zipfile import ZipFile
from pyproj import Transformer
from shapely.ops import transform
from shapely.geometry import Point
from pyproj import Geod
from shapely import LineString
from shapely.ops import substring
from shapely import MultiLineString
import warnings

warnings.filterwarnings("ignore", category=RuntimeWarning)

def hpdbCheck(raw_file_path):
    file = os.path.basename(raw_file_path)

    # Define the directory paths
    output_dir = 'Output'

    # Construct the full file paths
    output_file_path = os.path.join(output_dir, f"output_{file}")

    print("Processing = " + file)

    # Load the "kodepos.xlsx" Excel file into a pandas DataFrame
    kodepos_df = pd.read_excel('Reference/ZIP_Code.xlsx')

    # Load the "Mobile_Region_Cluster.xlsx" Excel file into a pandas DataFrame
    mobile_df = pd.read_excel('Reference/Mobile_Region_Cluster.xlsx')

    print(raw_file_path)

    # Load the file into a pandas DataFrame
    hpdb_df = pd.read_excel(raw_file_path)

    # Load City_Code.xlsx into a DataFrame
    city_code_df = pd.read_excel('Reference/City_Code.xlsx')

    #----------------------
    # City Code Lookup
    #----------------------
    # Merge the two DataFrames on the 'CITY' column
    merged_df = pd.merge(hpdb_df, city_code_df[['CITY', 'CITY_CODE']], on='CITY', how='left')

    # Update CITY_CODE in HPDB SAMPEL.xlsx with the values from merged_df
    hpdb_df['CITY_CODE'] = merged_df['CITY_CODE_y']

    #----------------------
    # Mobile Region and Cluster Lookup
    #----------------------
    # Create a copy of the CITY column in hpdb_df
    hpdb_df['CITY_original'] = hpdb_df['CITY']

    # Convert CITY columns to lowercase for case-insensitive comparison
    hpdb_df['CITY'] = hpdb_df['CITY'].str.lower()
    mobile_df['CITY'] = mobile_df['CITY'].str.lower()

    # Merge the DataFrames on 'CITY' column
    merged_df = pd.merge(hpdb_df, mobile_df, on='CITY', how='left', suffixes=('', '_MOBILE'))

    # Replace the values in MOBILE_REGION and MOBILE_CLUSTER with the corresponding values from REGION and CLUSTER
    merged_df['MOBILE_REGION'] = merged_df['REGION MOBILE'].combine_first(merged_df['MOBILE_REGION'])
    merged_df['MOBILE_CLUSTER'] = merged_df['CLUSTER MOBILE'].combine_first(merged_df['MOBILE_CLUSTER'])

    # Drop the extra 'REGION' and 'CLUSTER' columns
    merged_df.drop(['REGION MOBILE', 'CLUSTER MOBILE', 'PROVINCE', 'CITY.1'], axis=1, inplace=True)

    # Replace CITY column with original values
    merged_df['CITY'] = merged_df['CITY_original']

    # Drop the 'CITY_original' column
    merged_df.drop('CITY_original', axis=1, inplace=True)

    #----------------------
    # Null Filling
    #----------------------
    # Replace null values with '-'
    hpdb_filled_df = merged_df.fillna('-')

    # Remove double spaces from all columns
    hpdb_filled_df = hpdb_filled_df.replace(r'\s+', ' ', regex=True)

    # Remove '|' characters from PROJECT_NAME column
    hpdb_filled_df['PROJECT_NAME'] = hpdb_filled_df['PROJECT_NAME'].str.replace('|', '')

    # Create new column based on concatenated columns
    hpdb_filled_df['ADDRESS'] = hpdb_filled_df[['PREFIX_ADDRESS', 'STREET_NAME', 'HOUSE_NUMBER', 'BLOCK', 'FLOOR', 'RT', 'RW']].apply(lambda x: ' '.join(x.astype(str)), axis=1)

    # Save the filled DataFrame to a new Excel file
    hpdb_filled_df.to_excel('Temp/temp.xlsx', index=False)

    # Load HPDB SAMPEL.xlsx into a DataFrame
    hpdb_df = hpdb_filled_df

    wb = load_workbook('Temp/temp.xlsx')
    ws = wb.active

    #----------------------
    # List and Variables
    #----------------------
    # Initialize a PatternFill object for the red color
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Get the current year
    current_year = datetime.now().year

    # Define the lists of allowed values
    acquisition_class_values = ["HOME", "HOME - BIZ", "BIZ - HOME", "BIZ"]
    building_type_values = ['PERUMAHAN', 'RUKO', 'FASUM']
    ownership_values = ['PARTNERSHIP-LN', 'PARTNERSHIP-IF', 'PARTNERSHIP-TBG', 'OWN BUILT']
    vendor_name_values = ['LINKNET', 'IFORTE', 'TBG']
    prefix_address_values = ['JL.', 'GG.']

    # Define the lists of similar columns
    coordinate_col = ['FDT_LONGITUDE', 'FAT_LONGITUDE', 'BUILDING_LONGITUDE', 'FDT_LATITUDE', 'FAT_LATITUDE', 'BUILDING_LATITUDE']
    latitude_col = ['FDT_LATITUDE', 'FAT_LATITUDE', 'BUILDING_LATITUDE']
    longitude_col = ['FDT_LONGITUDE', 'FAT_LONGITUDE', 'BUILDING_LONGITUDE']
    rt_rw_col = ['RT','RW']

    # Convert the allowed values to lowercase
    acquisition_class_values_lower = [value.lower() for value in acquisition_class_values]
    building_type_values_lower = [value.lower() for value in building_type_values]
    ownership_values_lower = [value.lower() for value in ownership_values]
    vendor_name_values_lower = [value.lower() for value in vendor_name_values]
    prefix_address_values_lower = [value.lower() for value in prefix_address_values]

    # Define the regular expression patterns for longitude and latitude
    longitude_pattern = r'^\d{1,3}\.\d{6}$'
    latitude_pattern = r'^-?\d{1,2}\.\d{6}$'

    red_columns = []

    #----------------------
    # Data Converting
    #----------------------
    # Convert the datetime columns to datetime objects
    hpdb_df['PARTNER_RFS_DATE'] = pd.to_datetime(hpdb_df['PARTNER_RFS_DATE'], format='%m/%d/%Y', errors='coerce')
    hpdb_df['RFS_DATE'] = pd.to_datetime(hpdb_df['RFS_DATE'], format='%m/%d/%Y', errors='coerce')

    # Convert string columns to lowercase
    hpdb_df = hpdb_df.map(lambda x: x.lower() if isinstance(x, str) else x)
    kodepos_df = kodepos_df.map(lambda x: str(x).lower() if isinstance(x, str) else x)

    #-------------------------
    # LONGITUDE and LATITUDE
    #-------------------------
    # Replace all commas with dots and truncate decimals to 6 digits
    for col in coordinate_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                value = str(cell.value).replace(',', '.')  # Replace commas with dots
                if '.' in value:
                    integer_part, decimal_part = value.split('.')
                    decimal_part = decimal_part.ljust(6, '0')[:6]  # Ensure decimal part has exactly 6 digits
                    value = f"{integer_part}.{decimal_part}"
                cell.value = value

    # Define the latitude and longitude range
    latitude_range = (-12.0, 7.0)
    longitude_range = (94.0, 142.0)

    # Fill all existing rows with red color for longitude columns
    for col in longitude_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                value = float(cell.value)
                if not (longitude_range[0] <= value <= longitude_range[1]):
                    cell.fill = red_fill

    # Fill all existing rows with red color for latitude columns
    for col in latitude_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                value = float(cell.value)
                if not (latitude_range[0] <= value <= latitude_range[1]):
                    cell.fill = red_fill

    # Fill all existing rows with red color for RT and RW columns
    for col in rt_rw_col:
        for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
            for cell in row:
                if not re.match(r'^(-|\d+)$', str(cell.value)):
                    for cell in row:
                        cell.fill = red_fill
    
    #----------------------
    # Zip Code Checking
    #----------------------
    # Iterate over each row in the "HPDB SAMPEL" DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if there is a matching row in the "kodepos" DataFrame
        match = kodepos_df[
            (kodepos_df['REGION'] == row['REGION']) &
            # (kodepos_df['CITY'] == row['CITY']) &
            (kodepos_df['DISTRICT'] == row['DISTRICT']) &
            (kodepos_df['SUB_DISTRICT'] == row['SUB_DISTRICT']) &
            (kodepos_df['ZIP_CODE'].astype(str) == str(row['ZIP_CODE']))
        ]
        # If a match is found, mark the row in the Excel worksheet for coloring
        if match.empty:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('REGION') + 1, max_col=hpdb_df.columns.get_loc('ZIP_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # ACQUISITION_CLASS
    #----------------------
    # Fill all existing rows with red color if ACQUISITION_CLASS is not in allowed_values
    for index, row in hpdb_df.iterrows():
        if row['ACQUISITION_CLASS'].lower() not in acquisition_class_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # BUILDING_TYPE
    #----------------------
    # Iterate over each row in the "HPDB SAMPEL" DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if BUILDING_TYPE is one of the allowed values
        if row['BUILDING_TYPE'] not in building_type_values_lower:
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1, max_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # OWNERSHIP
    #----------------------
    # Check if any of the ownership values is in the OWNERSHIP field
    for index, row in hpdb_df.iterrows():
        found = False
        for value in ownership_values_lower:
            if value in row['OWNERSHIP']:
                found = True
                break
        if not found:
            # Fill the cell with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1, max_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1):
                for c in cell:
                    c.fill = red_fill
        elif row['OWNERSHIP'] != 'OWN BUILT' and row['VENDOR_NAME'] not in vendor_name_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1, max_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

    #----------------------
    # PREFIX_ADDRESS
    #----------------------
    for index, row in hpdb_df.iterrows():
        if row['PREFIX_ADDRESS'] not in prefix_address_values_lower:
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1):
                for c in cell:
                    c.fill = red_fill

    #--------------------------------------------------------
    # CLUSTER_NAME, CLUSTER_CODE, PROJECT_NAME, CITY_CODE
    #--------------------------------------------------------
    # Iterate over each row in the DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if the CLUSTER_NAME column exceeds 100 characters
        if len(str(row['CLUSTER_NAME'])) > 100 or row['CLUSTER_NAME'] == "-":
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the CLUSTER_CODE column exceeds 20 characters
        if len(str(row['CLUSTER_CODE'])) > 20 or row['CLUSTER_CODE'] == "-":
            # Mark the row in the Excel worksheet for coloring
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the length of PROJECT_NAME exceeds 100 characters
        if len(row['PROJECT_NAME']) > 100 or row['PROJECT_NAME'] == "-":
            # Fill the entire row with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1, max_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1):
                for c in cell:
                    c.fill = red_fill

        # Check if the length of PROJECT_NAME exceeds 100 characters
        if row['CITY_CODE'] == "-":
            # Fill the entire row with red color
            for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CITY_CODE') + 1, max_col=hpdb_df.columns.get_loc('CITY_CODE') + 1):
                for c in cell:
                    c.fill = red_fill

    #-------------
    # HOMEPASS_ID
    #-------------
    # Get the duplicated rows in the "HOMEPASS_ID" column
    duplicates = hpdb_df["HOMEPASS_ID"].duplicated()

    # Fill the duplicated rows with red color
    for index, duplicate in duplicates.items():
        if duplicate:
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('HOMEPASS_ID') + 1).fill = red_fill

    # Iterate over each row in the DataFrame
    for index, row in hpdb_df.iterrows():
        # Check if the PARTNER_RFS_DATE column is NULL or has an incorrect format
        if pd.isnull(row['PARTNER_RFS_DATE']) or \
            not isinstance(row['PARTNER_RFS_DATE'], pd.Timestamp) or \
            row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y') != row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y'):
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('PARTNER_RFS_DATE') + 1).fill = red_fill

        # Check if the RFS_DATE column is NULL or has an incorrect format
        if pd.isnull(row['RFS_DATE']) or \
            not isinstance(row['RFS_DATE'], pd.Timestamp) or \
            row['RFS_DATE'].strftime('%m/%d/%Y') != row['RFS_DATE'].strftime('%m/%d/%Y'):
            ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('RFS_DATE') + 1).fill = red_fill

    #-----------------------------
    # Address Duplicate Checking
    #-----------------------------
    duplicates = hpdb_filled_df["ADDRESS"].duplicated()

    # Fill the duplicated rows with red color
    for index, duplicate in duplicates.items():
        if duplicate:
            for row in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('RW') + 1):
                for cell in row:
                    cell.fill = red_fill

    # Drop the "ADDRESS" column
    ws.delete_cols(hpdb_filled_df.columns.get_loc("ADDRESS") + 1, amount=1)

    #-------------
    # Save
    #-------------
    # Save the modified Excel workbook
    wb.save(output_file_path)

    #-----------------------------
    # Summary
    #-----------------------------
    # Load the Excel workbook
    wb = load_workbook(output_file_path)
    ws = wb.active

    # Get the names of the columns from the first row
    hpdb_col = [cell.value for cell in ws[1]]

    # Iterate through each column
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        red_found = False
        # Check if any cell in the column has the red_fill pattern fill
        for cell in col:
            if cell.fill == red_fill:  # Check for red_fill pattern fill
                red_columns.append(hpdb_col[col_idx - 1])  # Append the column name
                red_found = True
                break
            if red_found:
                break  # No need to continue checking other cells in the column if one is already red

    # Create a DataFrame with column names from column_names
    hpdb_summary_df = pd.DataFrame(columns=hpdb_col)

    # Create a new row
    new_hpdb_row = {}
    for col_name in hpdb_col:
        if col_name in red_columns:
            new_hpdb_row[col_name] = "Revise"
        else:
            new_hpdb_row[col_name] = "OK"

    # Append the new row to the DataFrame
    hpdb_summary_df = hpdb_summary_df._append(new_hpdb_row, ignore_index=True)

    # Remove the raw file
    # os.remove(raw_file_path)

    print("Done")

    return pd.DataFrame(hpdb_summary_df)

def get_placemark(file_path):
  kmz = ZipFile(file_path, 'r')
  doc = kmz.open('doc.kml', 'r').read()
  kml_file = html.fromstring(doc)
  placemark_dict = {}
  for pm in kml_file.cssselect("Folder"):
    folder_name = pm.cssselect("name")[0].text_content()
    if folder_name not in placemark_dict:
      placemark_dict[folder_name] = pm.cssselect("Placemark")
    elif folder_name in placemark_dict:
      placemark_dict[folder_name + "_2"] = pm.cssselect("Placemark")
  return placemark_dict

# @title run this (just once)
def get_homepass_folder(placemark_dict):
  folder = []
  for hp_folder in ["HOME", "HOME-BIZ", "BIZ-HOME", "BIZ"]:
    if hp_folder in placemark_dict.keys():
      folder.append(hp_folder)
  return folder

def long_lat_mapping(coor):
  longitude = []
  latitude = []
  coordinates = coor
  for o, x in enumerate(coordinates):
    if o % 2 == 0:
      longitude.append(x)
    else:
      latitude.append(x)
  return longitude, latitude

def to_df(placemark_dict, parse_simple=False, mapping=False):
  data = []
  for pm in placemark_dict:
    for name in pm.cssselect("name"):
      name = name.text_content()
    for coord in pm.cssselect("coordinates"):
      coords = [float(i.replace("0 ", "")) for i in coord.text_content().strip().split(",") if len(i)>=1][:-1]
      if mapping:
        longitude, latitude = long_lat_mapping(coords)
        coords = list(zip(longitude, latitude))
    if parse_simple:
      for simple in pm.cssselect("SimpleData"):
        if simple.attrib['name'] == "FAT_CODE":
          FAT_CODE = simple.text_content()
        if simple.attrib["name"] == "HOMEPASS_ID":
          homepass_id = simple.text_content()
    if parse_simple:
      data.append([name, coords, FAT_CODE, homepass_id])
    else:
      data.append([name, coords])
  if parse_simple:
    df = pd.DataFrame(data=data, columns=["Name", "Coordinates", "FAT_CODE", "homepass_id"])
  else:
    df = pd.DataFrame(data=data, columns=["Name", "Coordinates"])
  return df

def make_circle_coord(longitude, latitude, radius=2):
  local_azimuthal_projection = "+proj=aeqd +R=6371000 +units=m +lat_0={} +lon_0={}".format(latitude, longitude)
  wgs84_to_aeqd = Transformer.from_proj('+proj=longlat +datum=WGS84 +no_defs', local_azimuthal_projection)
  aeqd_to_wgs84 = Transformer.from_proj(local_azimuthal_projection, '+proj=longlat +datum=WGS84 +no_defs')
  # Get polygon with lat lon coordinates
  point_transformed = Point(wgs84_to_aeqd.transform(longitude, latitude))

  buffer = point_transformed.buffer(radius)
  circle = transform(aeqd_to_wgs84.transform, buffer)
  return circle

def mapping_hp_to_pole(hp_df, pole_df):
  hp_df = hp_df
  geod = Geod(ellps="WGS84")
  pole_list = []
  for _, hp in hp_df.iterrows():
    hp_point = Point(hp["Coordinates"][0], hp["Coordinates"][1])
    distance_list = []
    for pole_index, pole in pole_df.iterrows():
      pole_point = Point(pole["Coordinates"][0], pole["Coordinates"][1])
      line = LineString([hp_point, pole_point])
      distance_list.append(geod.geometry_length(line))
    lowest_distance = min(distance_list)
    index_of_lowest_value = distance_list.index(lowest_distance)
    pole_list.append(index_of_lowest_value)
  hp_df["Pole_index"] = pole_list
  return hp_df

def check_pole_to_hp(placemark_dict, pole_df):
  hp_folder_name = get_homepass_folder(placemark_dict)
  pole_to_hp_35m = []
  for name in hp_folder_name:
    homepass_df = to_df(placemark_dict[name], parse_simple=True)
    geod = Geod(ellps="WGS84")
    for _, hp in homepass_df.iterrows():
      hp_point = Point(hp["Coordinates"][0], hp["Coordinates"][1])
      distance_list = []
      for _, pole in pole_df.iterrows():
        pole_point = Point(pole["Coordinates"][0], pole["Coordinates"][1])
        line = LineString([hp_point, pole_point])
        distance_list.append(geod.geometry_length(line))
      lowest_distance = min(distance_list)
      if lowest_distance > 35:
        pole_to_hp_35m.append("{} {} : {}".format(name, hp["Name"], ((hp["Coordinates"][0], hp["Coordinates"][1]))))   
    return pole_to_hp_35m

def check_fat_to_hp(placemark_dict, pole_df, fat_df, cable_df):
  fat_to_hp_150m = []
  hp_folder_name = get_homepass_folder(placemark_dict)
  try:
    sling_df = to_df(placemark_dict["SLING"], parse_simple=False, mapping=True)
  except:
    has_sling = False
  for name in hp_folder_name:
    homepass_df = to_df(placemark_dict[name], parse_simple=True)
    geod = Geod(ellps="WGS84")
    hp_df = mapping_hp_to_pole(homepass_df, pole_df)
    count_case1 = 0
    count_case2 = 0
    count_case3 = 0
    case3_name = []
    case3_coord = []
    count_case4 = 0
    case4_name = []
    case4_coord = []
    for _, hp in hp_df.iterrows():
      fat = fat_df[fat_df["Name"] == hp["FAT_CODE"]]
      if len(fat) == 0:
        fat_to_hp_150m["there is no fat {} in kmz file".format(hp["FAT_CODE"])] = hp["Coordinates"]
        continue
      pole_coordinates = pole_df["Coordinates"].iloc[hp["Pole_index"]]
      hp_point = Point(hp["Coordinates"][0], hp["Coordinates"][1])
      fat_point = Point(fat["Coordinates"].item()[0], fat["Coordinates"].item()[1])
      pole_point = Point(pole_coordinates[0], pole_coordinates[1])
      pole_area = make_circle_coord(pole_coordinates[0], pole_coordinates[1], radius=3)
      dropwire_line = LineString([hp_point, pole_point])
      dropwire_length = geod.geometry_length(dropwire_line)
      has_sling_line = False
      if has_sling:
        for _, sling in sling_df.iterrows():
          if len(sling["Coordinates"]) > 1:
            tmp_sling_line = LineString(sling["Coordinates"])
            if tmp_sling_line.intersects(pole_area):
              has_sling_line = tmp_sling_line.intersects(pole_area)
      tmp_distance = []
      for _, cable in cable_df.iterrows():
        tmp_cable_line = LineString(cable["Coordinates"])
        dist_in_cable = tmp_cable_line.line_locate_point(pole_point)
        point_in_cable = tmp_cable_line.line_interpolate_point(dist_in_cable)
        tmp_line = LineString([point_in_cable, pole_point])
        tmp_distance.append(geod.geometry_length(tmp_line))
      lowest_distance = min(tmp_distance)
      index_of_lowest_value = tmp_distance.index(lowest_distance)
      cable = cable_df.iloc[index_of_lowest_value]
      cable_line = LineString(cable["Coordinates"])
      pole_to_fat_line = substring(cable_line, start_dist=cable_line.line_locate_point(fat_point), end_dist=cable_line.line_locate_point(pole_point))
      pole_to_fat_length = geod.geometry_length(pole_to_fat_line)
      distance = dropwire_length + pole_to_fat_length
      if has_sling_line:
        pole_to_fat_line_air = LineString([pole_point, fat_point])
        pole_to_fat_air_length = geod.geometry_length(pole_to_fat_line_air)
        distance = dropwire_length + pole_to_fat_air_length
      if distance > 150:
        count_case2 += 1
        pole_to_fat_line_air = LineString([pole_point, fat_point])
        pole_to_fat_air_length = geod.geometry_length(pole_to_fat_line_air)
        distance_case2 = dropwire_length + pole_to_fat_air_length
        if distance_case2 <= 150:
          count_case3 += 1
          fat_to_hp_150m.append("{} {} : {}".format(name, hp["Name"], ((hp["Coordinates"][0], hp["Coordinates"][1]))))
        elif distance_case2 > 150:
          count_case4 += 1
          fat_to_hp_150m.append("{} {} : {}".format(name, hp["Name"], ((hp["Coordinates"][0], hp["Coordinates"][1]))))
  return fat_to_hp_150m

def is_fat_contain_pole(df_pole, df_fat):
  fat_pole = {}
  does_not_contain_pole = []
  for _, row_fat in df_fat.iterrows():
    circle = make_circle_coord(row_fat["Coordinates"][0], row_fat['Coordinates'][1], radius = 2)
    for _, row_pole in df_pole.iterrows():
      p = Point(row_pole["Coordinates"][0], row_pole["Coordinates"][1])
      condition = circle.contains(p)
      if condition:
        if row_fat["Name"] not in fat_pole.keys():
          fat_pole[row_fat["Name"]] = [row_pole["Name"]]
        else:
          fat_pole[row_fat["Name"]].append(row_pole["Name"])
      else:
        if row_fat["Name"] not in fat_pole.keys():
          fat_pole[row_fat["Name"]] = []
        else:
          continue
  for key, values in fat_pole.items():
    if len(values) == 1:
      continue
    else:
      does_not_contain_pole.append(key)
  return does_not_contain_pole

def is_fdt_contain_pole(df_pole, df_fdt):
  fdt_pole = {}
  does_not_contain_pole = []
  for _, row_fdt in df_fdt.iterrows():
    circle = make_circle_coord(row_fdt["Coordinates"][0], row_fdt['Coordinates'][1], radius = 2)
    for _, row_pole in df_pole.iterrows():
      p = Point(row_pole["Coordinates"][0], row_pole["Coordinates"][1])
      condition = circle.contains(p)
      if condition:
        if row_fdt["Name"] not in fdt_pole.keys():
          fdt_pole[row_fdt["Name"]] = [row_pole["Name"]]
        else:
          fdt_pole[row_fdt["Name"]].append(row_pole["Name"])
      else:
        if row_fdt["Name"] not in fdt_pole.keys():
          fdt_pole[row_fdt["Name"]] = []
        else:
          continue
  for key, values in fdt_pole.items():
    if len(values) == 1:
      continue
    else:
      does_not_contain_pole.append(key)
  return does_not_contain_pole

def check_duplicate_hpid(homepass_df):
  mask = homepass_df["homepass_id"].duplicated()
  duplicate = homepass_df[mask]
  for i, j in duplicate.iterrows():
    print("house number {} with homepass id {} has duplicate".format(j["Name"], j["homepass_id"]))

def check_distribution_cable_connect_to_pole(df_cable, df_pole):
  no_intersection_count = {}
  for _, pole in df_pole.iterrows():
    pole_area = make_circle_coord(pole["Coordinates"][0], pole["Coordinates"][1], radius=2)
    for _, cable in df_cable.iterrows():
      cable_line = LineString(cable['Coordinates'])
      intersect = cable_line.intersects(pole_area)
      if not intersect:
        if pole["Name"] not in no_intersection_count:
          no_intersection_count[pole["Name"]] = 1
        else:
          no_intersection_count[pole["Name"]] += 1
  cable_not_in_distribution = []
  for name, count in no_intersection_count.items():
    if count == len(df_cable):
      cable_not_in_distribution.append(name)
  filtered_df = df_pole[df_pole['Name'].isin(cable_not_in_distribution)]
  return filtered_df

def check_cable_distribution_has_sling(df_cable, df_pole, df_sling):
  pole_not_in_distribution_df = check_distribution_cable_connect_to_pole(df_cable, df_pole)
  no_intersection_count = {}
  for _, pole in pole_not_in_distribution_df.iterrows():
    pole_area = make_circle_coord(pole["Coordinates"][0], pole["Coordinates"][1], radius=2)
    for _, sling in df_sling.iterrows():
      if len(sling["Coordinates"]) > 1:
        sling_line = LineString(sling["Coordinates"])
        intersect = sling_line.intersects(pole_area)
      else:
        continue
      if not intersect:
        if pole["Name"] not in no_intersection_count:
          no_intersection_count[pole["Name"]] = 1
        else:
          no_intersection_count[pole["Name"]] += 1
  pole_not_in_sling = []
  for name, count in no_intersection_count.items():
    if count == len(df_sling):
      pole_not_in_sling.append(name)
  filtered_df = pole_not_in_distribution_df[pole_not_in_distribution_df['Name'].isin(pole_not_in_sling)]
  return filtered_df

def kmzCheck(file_path):   
    placemark_dict = get_placemark(file_path)
    try:
      pole_df = to_df(placemark_dict["POLE"], parse_simple=False)
    except:
      print("no folder POLE")
    try:
      fat_df = to_df(placemark_dict["FAT"], parse_simple=False)
    except:
      print("no folder FAT")
    try:
      fdt_df = to_df(placemark_dict["FDT"], parse_simple=False)
    except:
      print("no folder FDT")
    try:
      cable_df = to_df(placemark_dict["CABLE DISTRIBUTION"], parse_simple=False, mapping=True)
    except:
      print("no folder CABLE DISTRIBUTION")
    try:
      sling_df = to_df(placemark_dict["SLINGWIRE"], parse_simple=False, mapping=True)
    except:
      print("no folder SLINGWIRE")

    pole_to_hp = check_pole_to_hp(placemark_dict, pole_df)
    fat_to_hp = check_fat_to_hp(placemark_dict, pole_df, fat_df, cable_df)      
    fat_to_pole = is_fat_contain_pole(pole_df, fat_df)
    fdt_to_pole = is_fdt_contain_pole(pole_df, fdt_df)

    gilang_col = ["Result", 
    'Pole to FAT', 'Pole to FDT',
    'HP to pole 35m', 'HP to FAT 150m',]

    # Create a DataFrame with column names from column_names
    gilang_df = pd.DataFrame(columns=gilang_col)
    results = 0
    if len(pole_to_hp) == 0:
      pole_to_hp = "OK"
      results += 1
    if len(fat_to_hp) == 0:
      fat_to_hp = "OK"
      results += 1
    if len(fat_to_pole) == 0:
      fat_to_pole = "OK"
      results += 1
    if len(fdt_to_pole) == 0:
      fdt_to_pole = "OK"
      results += 1

    if results == 4:
      results = "OK"
    else:
      results = "REVISE"

    var_list = [results ,fat_to_pole, fdt_to_pole, pole_to_hp, fat_to_hp]
    

    new_row = {}
    for col_name, var in zip(gilang_col, var_list):
      new_row[col_name] = str(var)

    # Append the new row to the DataFrame
    gilang_df = gilang_df._append(new_row, ignore_index=True)

    return pd.DataFrame(gilang_df)

def main():
    log_columns = ['Cluster ID', 'Checking date', 'Checking Time']

    print("FORCE is Running...")

    clusters = os.listdir('Input')

    for cluster in clusters:
        input_dir = f"Input\{cluster}"
        summary_dir = f"Summary\{cluster}"
        kmz_file_path = os.path.join(input_dir, f"ABD - {cluster}.kmz")
        hpdb_file_path = os.path.join(input_dir, f"HPDB - {cluster}.xlsx")
        summary_file_path = os.path.join(summary_dir, f"Summary_{cluster}.xlsx")

        # Membuat direktori jika belum ada
        os.makedirs(summary_dir, exist_ok=True)

        print(hpdb_file_path)

        start_time = time.time()

        # hdpb_df = hpdbCheck(hpdb_file_path)
        # kmz_df = kmzCheck(kmz_file_path)

        # Jalankan kedua fungsi secara paralel
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future1 = executor.submit(hpdbCheck, hpdb_file_path)
            future2 = executor.submit(kmzCheck, kmz_file_path)

            # Ambil hasil kembali dari kedua fungsi
            hdpb_df = future1.result()
            kmz_df = future2.result()

        end_time = time.time()

        execution_time = end_time - start_time
        print("Execution time:", execution_time, "seconds")

        # Get current date
        checking_date = datetime.today().strftime('%Y-%m-%d')

        # Get current time
        checking_time = datetime.now().strftime('%H:%M:%S')

        # Create a DataFrame with the log_columns and the new row
        log_columns_df = pd.DataFrame(columns=log_columns)
        log_columns_df.loc[0] = [cluster, checking_date, checking_time]

        summary_df = pd.concat([log_columns_df, kmz_df, hdpb_df], axis=1, ignore_index=False)

        # Create temp_master.xlsx with headers from master_temp_df if it doesn't exist
        if not os.path.exists(summary_file_path):
            summary_df.to_excel(summary_file_path, index=False, engine='openpyxl')
        
        else:
            with pd.ExcelWriter(summary_file_path, 'openpyxl', mode='a',  if_sheet_exists="overlay") as writer:
                # fix line
                reader = pd.read_excel(summary_file_path, sheet_name="Sheet1")
                summary_df.to_excel(writer, sheet_name="Sheet1", index=False, engine='openpyxl', header=False, startrow=len(reader)+1)

if __name__ == "__main__":
    main()